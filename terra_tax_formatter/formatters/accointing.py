from openpyxl import load_workbook, Workbook
from io import BytesIO
from collections import OrderedDict
from .base import BaseFormatter

class Accointing(BaseFormatter):
    def __init__(self):
        super().__init__(
            headers=[
                "transactionType",
                "date",
                "inBuyAmount",
                "inBuyAsset",
                "outSellAmount",
                "outSellAsset",
                "feeAmount (optional)",
                "feeAsset (optional)",
                "classification (optional)",
                "operationId (optional)"
            ],
            headers_to_data_keys = {
                "date": "date",
                "operationId (optional)": "txhash",
                "inBuyAmount": "received quantity",
                "inBuyAsset": "received currency",
                "transactionType": "type"
            }
        )
        

        self.stake_tax_endpoint = "/accointing.xlsx"
        self.read_file_mode = "rb"

    def parse_workbook(self, wb):
        ws = wb[wb.sheetnames[0]]
        header = [cell.value for cell in ws[1]]

        values = []
        for row in list(ws.rows)[1:]:
            value = OrderedDict()
            for key, cell in zip(header, row):
                value[key] = cell.value
            values.append(value)
        
        return values

    def parse_file_data(self, file_obj):
        wb = load_workbook(file_obj)
        return self.parse_workbook(wb)

    def parse_response_data(self, file_obj):
        wb = load_workbook(BytesIO(file_obj.content))
        return self.parse_workbook(wb)

    def write_data(self, data, fname):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"

        ws1.append(self.headers)

        for row in data:
            ws1.append(list(row.values()))

        wb.save(fname)

    def get_unique_tx_hashes(self, data):
        return list(set([tx["operationId (optional)"] for tx in data]))

    def format_date(self, date_obj):
        return date_obj.strftime("%m/%d/%Y %H:%M:%S")
